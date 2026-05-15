import json
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from app.api.deps import get_db
from app.models.company import Company
from app.core.knowledge.chunker import DocumentChunker
from app.core.knowledge.embedder import EmbeddingService
from app.core.knowledge.weaviate_store import WeaviateStore
from app.core.knowledge.retriever import CompanyRetriever

router = APIRouter(prefix="/api/v1/knowledge", tags=["knowledge"])


class EntityCreate(BaseModel):
    name: str = Field(..., max_length=200)
    entity_type: str = Field(default="company", pattern="^(company|individual)$")
    industry: str | None = None
    scale: str | None = None
    description: str | None = None
    tags: list[str] = []


class SearchRequest(BaseModel):
    query: str = Field(..., min_length=1)


def _vectorize_and_store(company: Company):
    """Helper: chunk + embed + store a company's description into Weaviate."""
    if not company.description:
        return
    chunker = DocumentChunker()
    embedder = EmbeddingService()
    store = WeaviateStore()
    store.init_schema(embedder.dimension)
    chunks = chunker.split(company.description)
    embeddings = embedder.embed(chunks)
    tags_str = company.tags or "[]"
    store.insert_chunks(
        company_id=company.id,
        company_name=company.name,
        industry=company.industry or "",
        tags=tags_str,
        chunks=chunks,
        embeddings=embeddings,
    )


@router.post("/entities")
def create_entity(req: EntityCreate, db: Session = Depends(get_db)):
    """手动录入单个企业或个人"""
    db_entity = Company(
        name=req.name,
        entity_type=req.entity_type,
        industry=req.industry,
        scale=req.scale if req.entity_type == "company" else None,
        description=req.description,
        tags=json.dumps(req.tags, ensure_ascii=False),
    )
    db.add(db_entity)
    db.flush()
    _vectorize_and_store(db_entity)
    db.commit()
    return {"id": db_entity.id, "name": db_entity.name, "entity_type": db_entity.entity_type}


@router.get("/entities")
def list_entities(skip: int = 0, limit: int = 50, entity_type: str | None = None,
                  db: Session = Depends(get_db)):
    """分页查询已录入的实体列表"""
    q = db.query(Company)
    if entity_type:
        q = q.filter(Company.entity_type == entity_type)
    total = q.count()
    rows = q.order_by(Company.created_at.desc()).offset(skip).limit(limit).all()
    return {
        "total": total,
        "items": [
            {
                "id": r.id,
                "name": r.name,
                "entity_type": r.entity_type,
                "industry": r.industry,
                "scale": r.scale,
                "tags": json.loads(r.tags) if r.tags else [],
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
            for r in rows
        ],
    }


@router.post("/entities/upload")
def upload_entities(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """通过Excel批量上传企业和个人

    Excel表头（第一行）:
        name | entity_type | industry | scale | description | tags
    - entity_type: company 或 individual
    - tags: 逗号分隔
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "仅支持 .xlsx / .xls 格式")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "服务器缺少 openpyxl 依赖")

    content = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    required = {"name", "entity_type", "industry", "scale", "description", "tags"}
    if not required.issubset(set(headers)):
        raise HTTPException(400, f"Excel表头必须包含: {required}")

    created = []
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = dict(zip(headers, row))
        name = str(data.get("name") or "").strip()
        if not name:
            continue

        entity_type = str(data.get("entity_type") or "company").strip()
        if entity_type not in ("company", "individual"):
            errors.append({"row": row_idx, "error": f"entity_type 无效: {entity_type}"})
            continue

        tags_str = str(data.get("tags") or "")
        tags = [t.strip() for t in tags_str.split(",") if t.strip()]

        try:
            db_entity = Company(
                name=name,
                entity_type=entity_type,
                industry=str(data.get("industry") or "").strip() or None,
                scale=str(data.get("scale") or "").strip() if entity_type == "company" else None,
                description=str(data.get("description") or "").strip() or None,
                tags=json.dumps(tags, ensure_ascii=False),
            )
            db.add(db_entity)
            db.flush()
            _vectorize_and_store(db_entity)
            created.append({"id": db_entity.id, "name": name, "entity_type": entity_type})
        except Exception as e:
            errors.append({"row": row_idx, "error": str(e)})

    db.commit()
    return {"created": len(created), "errors": len(errors), "items": created, "error_details": errors[:10]}


@router.post("/search")
def search_entities(req: SearchRequest):
    retriever = CompanyRetriever()
    results = retriever.retrieve(req.query)
    return results
